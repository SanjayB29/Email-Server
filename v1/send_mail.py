import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def send_mail(sender, to, content):
    """
    Appends mail to the sender's and receiver's sheets in the Excel file.
    """
    # Load the workbook and get the Users sheet to check if recipient exists
    workbook = load_workbook('email_server.xlsx')
    users_df = pd.read_excel('email_server.xlsx', sheet_name='Users')

    if to not in users_df['EmailID'].values:
        return False, "Recipient email not found."

    mail_id = generate_mail_id()  # Generate a unique mail ID for this email

    # Data to append
    sender_data = {'To': to, 'Contents': content, 'MailID': mail_id}
    receiver_data = {'From': sender, 'Contents': content, 'MailID': mail_id}

    # Append data to sender's sheet
    if sender in workbook.sheetnames:
        append_to_sheet(workbook, sender, sender_data, 'sent')
    else:
        return False, "Sender sheet not found."

    # Append data to receiver's sheet
    if to in users_df['Username'].values:
        recipient_username = users_df[users_df['EmailID'] == to]['Username'].values[0]
        append_to_sheet(workbook, recipient_username, receiver_data, 'received')
    else:
        return False, "Recipient username not found."

    workbook.save('email_server.xlsx')
    workbook.close()
    return True, "Mail sent successfully."

def append_to_sheet(workbook, sheet_name, data, mail_type):
    """
    Appends data to a specified sheet within the workbook.
    """
    sheet = workbook[sheet_name]
    if mail_type == 'sent':
        # Find the first empty row in sent columns
        row = find_first_empty_row(sheet, 'D')
        sheet[f'D{row}'] = data['To']
        sheet[f'E{row}'] = data['Contents']
        sheet[f'F{row}'] = data['MailID']
    elif mail_type == 'received':
        # Find the first empty row in received columns
        row = find_first_empty_row(sheet, 'A')
        sheet[f'A{row}'] = data['From']
        sheet[f'B{row}'] = data['Contents']
        sheet[f'C{row}'] = data['MailID']
    else:
        raise ValueError("Invalid mail type specified.")

def find_first_empty_row(sheet, column):
    """
    Finds the first empty row in a specific column.
    """
    row = 1
    while row <= sheet.max_row:
        if not sheet[f'{column}{row}'].value:
            break
        row += 1
    return row if row <= sheet.max_row else row + 1

def generate_mail_id():
    """
    Generates a unique mail ID. Implement this function based on your requirements.
    """
    # Placeholder implementation. Customize as needed.
    return pd.Timestamp.now().strftime('%Y%m%d%H%M%S%f')

def send_mail_page():
    st.title("Send Mail")
    sender_email = st.session_state['email']  # Assuming sender's email is stored in session_state
    to = st.text_input("To Email:")
    content = st.text_area("Content:")
    
    if st.button("Send"):
        success, message = send_mail(sender_email, to, content)
        if success:
            st.success(message)
        else:
            st.error(message)
    
    if st.button("Back"):
        st.session_state['current_page'] = 'main_page'  # Ensure this matches your app's page routing logic
        st.experimental_rerun()

# Example usage in your Streamlit app
if __name__ == '__main__':
    if 'current_page' not in st.session_state:
        st.session_state['current_page'] = 'send_mail_page'  # Default page for demonstration
    
    # Assuming you have a way to set 'email' in session_state upon user login
    st.session_state['email'] = 'sender@example.com'  # Placeholder email

    if st.session_state['current_page'] == 'send_mail_page':
        send_mail_page()
