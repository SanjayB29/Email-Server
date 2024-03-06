import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def signup(username, email, password):
    """
    Registers a new user by adding their credentials to the 'Users' sheet.
    It also creates a new sheet for the user to store emails.
    """
    workbook_path = 'email_server.xlsx'
    
    # Try to load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(workbook_path)
        users_df = pd.read_excel(workbook_path, sheet_name='Users')
    except (FileNotFoundError, InvalidFileException, KeyError):
        # Initialize workbook and users DataFrame if not existent
        workbook = load_workbook()
        workbook.save(workbook_path)
        workbook.close()
        users_df = pd.DataFrame(columns=['Username', 'EmailID', 'Password'])
    
    # Check if username or email already exists
    if not users_df.empty and (username in users_df['Username'].values or email in users_df['EmailID'].values):
        return False, "Username or Email already exists."
    
    # Add new user to the dataframe
    new_user = pd.DataFrame([[username, email, password]], columns=['Username', 'EmailID', 'Password'])
    users_df = pd.concat([users_df, new_user], ignore_index=True)
    
    # Save updated users list back to Excel
    with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        users_df.to_excel(writer, sheet_name='Users', index=False)
    
    # Create a new sheet for the user to store emails, if it doesn't already exist
    workbook = load_workbook(workbook_path)
    if username not in workbook.sheetnames:
        workbook.create_sheet(title=username)
        worksheet = workbook[username]
        worksheet.append(['From', 'To', 'Contents', 'MailID'])  # Adjust columns as needed
    workbook.save(workbook_path)
    workbook.close()
    
    return True, "Signup successful."

def login(email, password):
    """
    Authenticates a user by matching the email and password against the 'Users' sheet.
    """
    workbook_path = 'email_server.xlsx'
    
    # Attempt to load the 'Users' sheet to find the user
    try:
        users_df = pd.read_excel(workbook_path, sheet_name='Users')
    except (FileNotFoundError, InvalidFileException, KeyError):
        return False, None, "User database not found. Please initialize the database first."
    
    # Check for matching email and password
    user_row = users_df.loc[(users_df['EmailID'] == email) & (users_df['Password'] == password)]
    if not user_row.empty:
        return True, user_row.iloc[0]['Username'], "Login successful."
    else:
        return False, None, "Invalid Email or Password."
